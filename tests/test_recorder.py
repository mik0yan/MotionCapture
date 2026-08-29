from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from motion_capture.models import PoseSample
from motion_capture.recorder import CSVRecorder, TrajectoryRecorder


def test_recorder_writes_pose(tmp_path: Path) -> None:
    recorder = CSVRecorder(tmp_path)
    path = recorder.start()
    recorder.write((PoseSample("test", "tool", 7, (1.0, 2.0, 3.0), np.eye(3), 0.9),))
    recorder.stop()
    text = path.read_text(encoding="utf-8-sig")
    assert "timestamp_utc,source,tool_id" in text
    assert ",test,tool,7,1,0.900000,1.000000,2.000000,3.000000," in text


def test_csv_recorder_can_write_direction_cosine_matrix(tmp_path: Path) -> None:
    recorder = TrajectoryRecorder(tmp_path, file_format="csv", orientation="matrix")
    path = recorder.start()
    recorder.write((PoseSample("ndi", "01", 8, (4.0, 5.0, 6.0), np.eye(3), 0.8),))
    recorder.stop()

    lines = path.read_text(encoding="utf-8-sig").splitlines()
    assert "r11,r12,r13,r21,r22,r23,r31,r32,r33" in lines[0]
    assert "qw" not in lines[0]
    assert lines[1].endswith(
        "1.000000000,0.000000000,0.000000000,0.000000000,1.000000000,"
        "0.000000000,0.000000000,0.000000000,1.000000000"
    )


def test_xlsx_recorder_writes_numeric_quaternion_columns(tmp_path: Path) -> None:
    recorder = TrajectoryRecorder(tmp_path, file_format="xlsx", orientation="quaternion")
    path = recorder.start()
    recorder.write((PoseSample("ndi", "02", 9, (7.0, 8.0, 9.0), np.eye(3), 0.95),))
    recorder.stop()

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["trajectory"].iter_rows(values_only=True))
    workbook.close()
    assert rows[0][-4:] == ("qw", "qx", "qy", "qz")
    assert rows[1][1:4] == ("ndi", "02", 9)
    assert rows[1][-4:] == (1.0, 0.0, 0.0, 0.0)


def test_xlsx_recorder_can_write_direction_cosine_matrix(tmp_path: Path) -> None:
    recorder = TrajectoryRecorder(tmp_path, file_format="xlsx", orientation="matrix")
    path = recorder.start()
    recorder.write((PoseSample("ndi", "03", 10, (1.0, 2.0, 3.0), np.eye(3), 1.0),))
    recorder.stop()

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["trajectory"].iter_rows(values_only=True))
    workbook.close()
    assert rows[0][-9:] == ("r11", "r12", "r13", "r21", "r22", "r23", "r31", "r32", "r33")
    assert rows[1][-9:] == (1.0, 0.0, 0.0, 0.0, 1.0, 0.0, 0.0, 0.0, 1.0)


def test_xlsx_recorder_writes_realsense_position_without_orientation(tmp_path: Path) -> None:
    recorder = TrajectoryRecorder(tmp_path, file_format="xlsx", orientation="position")
    path = recorder.start()
    recorder.write(
        (PoseSample("realsense", "apriltag_board", 11, (12.5, -4.25, 600.0), np.eye(3), 0.75),)
    )
    recorder.stop()

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook["trajectory"].iter_rows(values_only=True))
    workbook.close()
    assert rows[0] == TrajectoryRecorder.BASE_HEADER
    assert rows[1][1:4] == ("realsense", "apriltag_board", 11)
    assert rows[1][-3:] == (12.5, -4.25, 600.0)
